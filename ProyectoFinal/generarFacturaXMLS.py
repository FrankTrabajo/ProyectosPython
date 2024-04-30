import sqlite3
import openpyxl
from datetime import datetime
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import win32com.client
import os

class FacturaExcel():

    def __init__(self, numeroFac):
        conexion = sqlite3.connect('database/usuarios.db')
        cursor = conexion.cursor()

        cursor.execute('SELECT * FROM facturas WHERE numeroFactura = ?', (numeroFac,))
        cliente = cursor.fetchone()  # Utilizamos fetchone() para obtener una sola fila
        habitacion = cliente[9]
        cursor.execute('SELECT precio FROM Habitacion WHERE numeroHab = ?', (habitacion,))
        precioHab = cursor.fetchall()
        precioHab = precioHab[0]

        if cliente is None:
            print(f"No se encontró ninguna factura con el número {numeroFac}.")
            return

        workbook = openpyxl.Workbook()
        hojaFactura = workbook.active





        # Insertar datos en el Excel
        datos_cliente = cliente[:10]  # Los primeros 10 elementos son los datos del cliente
        columnas = ["Nombre: ", "NIF: ", "Direccion: ", "CP: ", "Localidad: ", "Pais: ", "Personas: ", "Llegada: ", "Salida: ", "Habitacion: "]
        fila = 4
        for columna, valor in zip(columnas, datos_cliente):
            celda_columna = hojaFactura.cell(column=1, row=fila, value=columna)
            celda_valor = hojaFactura.cell(column=2, row=fila, value=valor)
            celda_valor.alignment = Alignment(horizontal='left')
            fila += 1


        fecha = datetime.now()
        año = fecha.year
        # Insertar número de factura
        celda_columna = hojaFactura.cell(column=1, row=1, value="Numero de Factura: ")
        celda_valor = hojaFactura.cell(column=2, row=1, value=str(numeroFac) + "/" + str(año))
        celda_valor.alignment = Alignment(horizontal='left')

        # Insertar precio
        celda_columna = hojaFactura.cell(column=1, row= fila, value="Precio de la habitacion: ")
        celda_valor = hojaFactura.cell(column=2, row=fila, value=str(precioHab[0])+"€")
        celda_columna = hojaFactura.cell(column=1, row= fila+1, value="IVA 10% Precio total: ")
        celda_valor = hojaFactura.cell(column=2, row=fila+1, value=str(cliente[11])+"€")
        celda_valor.alignment = Alignment(horizontal='left')

        nombre_cliente = cliente[0].replace(" ", "")
        nombre_archivo = f"facturas_{nombre_cliente}.xlsx"


        for columna  in hojaFactura.columns:
            hojaFactura.column_dimensions[columna[0].column_letter].width = 25

        # Guardar el archivo Excel
        workbook.save("facturas/" + nombre_archivo)

        # Imprime el archivo Excel
        xl_app = win32com.client.Dispatch("Excel.Application")
        xl_workbook = xl_app.Workbooks.Open(os.path.abspath("facturas/" + nombre_archivo))
        xl_worksheet = xl_workbook.Sheets[0]
        xl_worksheet.PrintOut()
        xl_workbook.Close(SaveChanges=False)
        xl_app.Quit()

