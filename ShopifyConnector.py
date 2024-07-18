import shopify
import requests
from openpyxl import Workbook
from openpyxl.styles import Font


#Configuracion de la URL de la tienda, la clave de la API y la contraseña
shop_url = "#url_de_tu_web.myshopify.com"
access_token = "#access_tokken_de_tu_web_de_la_aplicacion_creada_en_shopify"
api_version = "#api_version_de_la_aplicacion_creada_en_shopify"

api_session = shopify.Session(shop_url, api_version, access_token)
shopify.ShopifyResource.activate_session(api_session)

api_url = f"https://{shop_url}/admin/api/{api_version}/orders.json?status=any" #  =>  En este caso lo que se pide son los pedidos de nuestro shopify

response = requests.get(api_url, headers={"X-Shopify-Access-Token": access_token})
orders_json = response.json()

# Abrimos un documento TXT para guardar los datos en formato JSON
# Así es mas facil leer los datos que queremos guardar
with open("pedido.txt", "w") as file:
    file.write(str(orders_json))

# TODOS TODOS
# Almacena todos los datos de nuestro shopify en este caso (Pedidos)

headers = {
    "Content-Type": "application/json",
    "X-Shopify-Access-Token": access_token
}

# Definismos un metodo para ejecutar

def get_all_orders(url, headers):
    # Creamos una lista de pedidos
    orders = []
    while url:
        response = requests.get(url, headers=headers)
        # Si la conexion es estable (200 ó OK) almaceno los datos en nuestra lista
        if response.status_code == 200:
            data = response.json()
            orders.extend(data.get('orders', []))
            
            # Verificar si hay más páginas
            if 'next' in response.links:
                url = response.links['next']['url']
            else:
                url = None
        else:
            print(f"Error al obtener los pedidos: {response.status_code}")
            print(response.text)
            break
    return orders

# Obtener todos los pedidos
all_orders = get_all_orders(api_url, headers)

# Guardar pedidos en un archivo de texto
with open("pedido.txt", "w") as file:
    file.write(str(all_orders))

# Guardar pedidos en un archivo Excel
# Estos son unos de los campos que frecuentemente piden las empresas
titulos = ["Name", "Email", "Paid at", "Currency", "Subtotal", "Shipping", "Taxes", "Total", 
           "Lineitem quantity", "Lineitem name", "Lineitem price", "Billing name", "Billing street", 
           "Billing city", "Billing zip", "Billing province", "Billing country", "Billing phone", 
           "Payment method", "SKU"]

# Abrimos nuestra hoja donde vamos a almacenar los datos en nuestro Excel
wb = Workbook()
ws = wb.active
bold_font = Font(bold=True)

# Copiará en la primera fila los titulos para tener así las columnas ordenadas
for i in range(len(titulos)):
    cell = ws.cell(row=1, column=i+1, value=titulos[i])
    cell.font = bold_font

# Ahora buscamos en nuestra lista los titulos a guardar
# ¿Cómo sabemos que nombre tienen los titulos a guardar y si están o no dentro de algun subtipo?
# En shopify dev tenemos ejemplos de los JSON que se generan cuando realizamos alguna petición
for i, order in enumerate(all_orders):
    nameOrder = order.get("name", None)
    email = order.get("email", None)
    paid_at = order.get("created_at", None)
    currency = order.get("currency", None)
    current_subtotal_price = order.get("current_subtotal_price", None)

    total_shipping_price = None
    total_shipping_price_set = order.get("total_shipping_price_set", None)
    if total_shipping_price_set:
        shop_money = total_shipping_price_set.get("shop_money", None)
        if shop_money:
            total_shipping_price = shop_money.get("amount", None)

    total_tax = order.get("total_tax", None)
    current_total_price = order.get("current_total_price", None)
    
    lineitem_quantity = None
    lineitem_name = None
    lineitem_price = None
    line_items = order.get("line_items", [])
    if line_items:
        line_item = line_items[0]  # Consideramos solo el primer artículo de línea
        lineitem_quantity = line_item.get("quantity", None)
        lineitem_name = line_item.get("name", None)
        lineitem_price = line_item.get("price", None)
        lineitem_sku = line_item.get("sku", None)
        lineitem_fulfillment_status = line_item.get("fulfillment_status", None)

    customer = order.get("customer", None)
    name = None
    street = None
    city = None
    zip_code = None
    province = None
    country = None
    phone = None
    if customer:
        first_name = customer.get("first_name", None)
        last_name = customer.get("last_name", None)
        name = f"{first_name} {last_name}"

        default_address = customer.get("default_address", None)
        if default_address:
            street = default_address.get("address1", None)
            city = default_address.get("city", None)
            zip_code = default_address.get("zip", None)
            province = default_address.get("province", None)
            country = default_address.get("country_code", None)
            phone = default_address.get("phone", None)

    payment_gateway_names = order.get("payment_gateway_names", [])
    payment_method = payment_gateway_names[0] if payment_gateway_names else None

    # En este caso solo estamos almacenando esos epidos que aún no han sido realizados y están en estado pendientes
    # Si queremos almacenarlos todos simplemente quitamos el bucle If

    if str(lineitem_fulfillment_status) == 'None':
        ws.cell(row=i+2, column=1, value=nameOrder)
        ws.cell(row=i+2, column=2, value=email)
        ws.cell(row=i+2, column=3, value=paid_at)
        ws.cell(row=i+2, column=4, value=currency)
        ws.cell(row=i+2, column=5, value=current_subtotal_price)
        ws.cell(row=i+2, column=6, value=total_shipping_price)
        ws.cell(row=i+2, column=7, value=total_tax)
        ws.cell(row=i+2, column=8, value=current_total_price)
        ws.cell(row=i+2, column=9, value=lineitem_quantity)
        ws.cell(row=i+2, column=10, value=lineitem_name)
        ws.cell(row=i+2, column=11, value=lineitem_price)
        ws.cell(row=i+2, column=12, value=name)
        ws.cell(row=i+2, column=13, value=street)
        ws.cell(row=i+2, column=14, value=city)
        ws.cell(row=i+2, column=15, value=zip_code)
        ws.cell(row=i+2, column=16, value=province)
        ws.cell(row=i+2, column=17, value=country)
        ws.cell(row=i+2, column=18, value=phone)
        ws.cell(row=i+2, column=19, value=payment_method)
        ws.cell(row=i+2, column=20, value=lineitem_sku)
        ws.cell(row=i+2, column=21, value=str(lineitem_fulfillment_status))

# Siempre que trabajamos con archivos Excel debemos guardar los cambios

wb.save("Pedidos.xlsx")

# Cuando la aplicación ya ha acabado se muestra este mensaje

print("Pedidos guardados exitosamente en pedido.txt y Pedidos.xlsx")

